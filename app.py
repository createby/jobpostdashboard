from flask import Flask, request, jsonify, send_file, render_template
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json
import os
import io
from generate_pdf import generate_company_pdf

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "파일 없음"}), 400
    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    companies = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if not any(row):
            continue
        company = {}
        for j, val in enumerate(row):
            if j < len(headers):
                company[headers[j]] = str(val).strip() if val is not None else ""
        companies.append(company)
    return jsonify({"headers": headers, "companies": companies})

def parse_hire(data):
    if isinstance(data.get("채용계획"), str):
        try:
            data["채용계획"] = json.loads(data["채용계획"])
        except:
            data["채용계획"] = []
    return data

@app.route("/generate_pdf", methods=["POST"])
def generate_pdf_route():
    data = parse_hire(request.json)
    filename = f"{data.get('부스번호', 'X')}_{data.get('기업명', 'company')}.pdf"
    output_path = os.path.join(OUTPUT_FOLDER, filename)
    generate_company_pdf(data, output_path)
    return send_file(output_path, as_attachment=True, download_name=filename)

@app.route("/preview_pdf", methods=["POST"])
def preview_pdf_route():
    data = parse_hire(request.json)
    filename = f"{data.get('부스번호', 'X')}_{data.get('기업명', 'company')}.pdf"
    output_path = os.path.join(OUTPUT_FOLDER, filename)
    generate_company_pdf(data, output_path)
    return send_file(output_path, as_attachment=False, mimetype="application/pdf")

@app.route("/sample_excel")
def sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "확정기업"

    headers = [
        "번호", "부스번호", "ZONE", "기업명", "업종", "사업내용", "회사주소",
        "태그", "배경색", "배경색연한", "qr_url", "기업소개",
        "채용계획", "담당업무", "자격요건", "급여조건", "근무지역", "근무시간", "복리후생"
    ]

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="4CAF50")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) * 2.5, 15)

    # 샘플 행
    sample = [
        "1", "A01", "대·중견기업ZONE", "Adecco Korea", "서비스업", "HR컨설팅, 헤드헌팅, 스태핑",
        "서울시 송파구 백제고분로 69", "면접,상담", "#4CAF50", "#E8F5E9", "https://adecco.co.kr",
        "Adecco Group은 스위스 취리히에 본사를 두고 전세계 60개국가에서 HR Solution을 제공하는 글로벌 500대기업입니다.",
        '[["HR Consultant & Recruiter", "경력(1년이상)", "정규직", "대학교(4년) 졸업", "3명"]]',
        "고객사 채용 관리, 면접일정 조율, 계약관리, 근태관리",
        "HR산업 관심자 / 동종업계 경력자(1년 이상) / 영어 가능자",
        "회사 내규", "서울", "주 5일(월~금) 8시간",
        "연차휴가, 명절선물, 생일휴가, 탄력근무제, 재택근무(주1회)"
    ]
    ws.append(sample)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="채용박람회_샘플양식.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)
